from user_agents import user_agents
from openpyxl import Workbook, load_workbook
import json
import os
import random
import requests
import csv
import time
import output_format
import analysis


bad_link_prefixes = ["/search", "q=", "/?",
                     "/advanced_search"]

bad_locations = ["facebook", "instagram",
                 "linkedin", "twitter", "ratemyprofessors",
                 "coursicle", "youtube", "amazon",
                 ".doc", ".pdf", "wiki", "imgres"]


def build_output_file(file_path, header, log):
    """
    This builds the sheet, adds the relevant column names,
    then openpyxl.load_workbook() will be called to append
    new researchers as they are scraped and processed.

    file_path: the path of the excel file that is to be built
    header: the header of the new file
    log: a pyqtSignal(str) that will cause the GUI to be updated if the build
     fails
    return:
     True: if the build is good
     False: if the build is not. This will trigger the process ending and
      an error dialog box being displayed
    """

    wb = Workbook()
    ws = wb.active
    ws.append(header)
    # If the file cannot be built, that likely means that one of the same
    #  name exists elsewhere.
    try:
        wb.save(file_path)
        return True
    except PermissionError:
        log.emit("<br><br><b>AN ERROR WAS ENCOUNTERED WHEN ACCESSING THE" \
                 " OUTPUT EXCEL FILE. THIS LIKELY MEANS THAT IT IS OPEN" \
                 " SOMEWHERE. STOPING SCRAPING.</b>")
        return False


def read_csv(file_path, table, log):
    """
    Reads the csv at file_path and loads each researcher as a dict into the
    list to_search.

    file_path: the path of the csv that was inputted by the user.
    table: a pyqtSignal(str) that updates the table tab
    return: a list of dicts, where each dict contains
     'headers': a list containing all strings in the header
     'name': the person's name
     'institution': the person's institution
     plus any other column in the input csv
    """

    to_search = []
    for_table = ""
    with open(file_path, 'r') as f:
        reader = csv.reader(f)
        header = next(reader)

        for person in reader:
            person_data = {'data_from_csv': {header[i].lower() : person[i] \
                              for i in range(0,len(header))}}


            to_search.append(person_data)

            # there must be categories for name and institution in the input
            #  csv, otherwise the gui will show an error
            try:
                person_data['name'] = person_data['data_from_csv']['name']
                person_data['institution'] = person_data['data_from_csv']['institution']
            except KeyError:
                log.emit("<br><br> It looks like you did not include the" \
                         " name and/or the institution in the input csv.<br> " \
                         " Scraping is not possible without these, halting the" \
                         " process.")
                return False
    
    # the for_table string is built in this way to make parsing done
    #  for the table more easy.
    for_table +=  person_data['name']+ "," \
        + person_data['institution'] + "*"
    table.emit(for_table)
    return to_search


def get_links(person, sites, google_api_key, cse_id, log):
    """
    Gets relevant links from the first page of a google search for some person.

    person: a dict containing all pertinent information on some subject. Built
     by read_csv()
    sites: useful sites that the user inputted via the GUI. They will be
     appended to the base query. eg
      'https://www.google.com/search?q=Santa Claus North Pole'
      'https://www.google.com/search?q=Santa Claus North Pole NORAD Tracks Santa'
      Could be queries used for the individual Santa Claus if the only useful
      site that was inputted was NORAD Tracks Santa.
    agent: the agent used for searching
    log: a pyqtSignal(str) which emits useful information to the GUI log.
    return: all appropriate links found for some individual. See comments for
     a definition of appropriate
    """

    query = person['name'] + " " + person['institution']

    links = []

    # google's programmable search is used, to avoid getting locked out
    search_url = "https://www.googleapis.com/customsearch/v1?key=" \
        f"{google_api_key}&cx={cse_id}&q={query}"
    # creating a list of search terms to use
    all_search = [search_url + " " + site for site in sites]
    all_search = [search_url] + all_search

    # informs the user of the search queries. They are outputted as hyperlinks
    log.emit("<b>Initial search links used:</b><br>")
    to_log = [f"<a href=\"{site}\">{site}</a>" for site in all_search]
    log.emit("<br>".join(to_log) + "<br><br>")

    # gets the google search page, in bytes for each in all_search
    raw_links = []
    for search in all_search:
        response = requests.get(search)
        data = json.loads(response.text)

        if 'error' in data:
            log.emit("<br><br>There was an error while using Google's Programmable" \
                     f" Search:<br>{data['error']['message']}")
            continue
        elif 'items' not in data:
            continue

        raw_links += [item['link'] for item in data['items']]


    # filter links that do not contain "google.com" or start with the
    #  prefixes defined in bad_link_prefixes
    filtered_links = [link for link in raw_links
                      if not any(link.startswith(prefix)
                                 or link.find('google.com') > 0
                                 for prefix in bad_link_prefixes)]

    # filter links that don't contain searches
    filtered_links = [link for link in filtered_links
                      if not any(link.find(search) > -1
                                 for search in bad_locations)]

    # only grab the relevent part of the link if it includes more in it
    links += [link.split("/url?q=")[-1].split("&sa")[0]
              for link in filtered_links]
    links = [link for link in links if "/search" not in link]

    # finally, choosing good links based on our parameters:
    #  if any of the user-specified good sites are present in the link
    #  if the person's first and last name are present in the link
    #  if the person's institution is present in the link
    # This logic is heavily dependent on the quality of google's search
    #  algorithm, there are obviously a lot of holes and chances for bad data
    #  to slip in. However, this system works.
    full_name = person['name'].lower().split(" ")
    first_name = full_name[0]
    last_name = full_name[-1]
    institution = person['institution']
    if "researchgate" in sites:
        del sites[sites.index("researchgate")]
    links = [link for link in map(lambda x: x.lower(), set(links))
             if any(element in link for element in sites)
             or (first_name in link and last_name in link)
             or institution.lower() in link
             or ("researchgate" in link
                 and first_name in link
                 and last_name in link)]
    return links


def write_to_excel(output_path, person, log):
    """
    Writes each person to excel by using the package openpyxl.

    How it works is it checks the pre-written Excel sheet for header names (see
    build_output_file() and finds the corresponding key in the person dict
    that is passed to this funciton. The value for that key is then put into
    a list of length equal to the number of needed columns at the index corresponding
    to the column with the same name. After all of the keys and columns are matched
    up, the list is appended to the workbook and the workbook is closed.

    In this way, only needed columns are saved, and the workbook is updated
    after each person is scraped.

    output_path: the path of the Excel file to write to
    person: a dict containing all relevant and scraped information on one
     person
    log: a pyqtSygnal(str) object that is used to update the GUI's log
     with pertinent information
    """

    wb = load_workbook(output_path, data_only=True)
    ws = wb.active
    # getting the column names. this is a list of cells, not of strings 
    header = ws[1]
        
    # if there was no relevent data found for some person, then 'NONE'
    #  is written into the Excel cell.
    to_write = [""]*len(header)

    # outer loop cycles through the keys in the person's output. keys are
    #  checked to see if they exist as headers in the output and if they are
    #  then the value corresponding to that key is saved into the appropriate
    #  index in to_write
    for key in person['output']:
        # the inner loop cycles through each header object from the Excel sheet,
        #  if a match is found between the strings of key and h.internal_value,
        #  then the output is saved in that header's corresponding index on
        #  to_write
        for h in header:
            # gets the column integer of this header cell
            col = h.column

            # 'relevant links' is a reserved column; this will have the links
            #  found during scraping placed into it.
            if h.internal_value.lower() == "relevant links":
                to_write[col - 1] = "\n".join(person['links used'])
                continue 

            data = person['output'][key]
            # splitting the data string like this, then removing each first
            #  blank string item is done for formatting purposes.
            items = data.split("\n")
            while len(items) > 0 and items[0] == "":
                del items[0]

            # using a set here is done to try and eliminate any repetition of
            #  information
            unique = "\n".join(set(items))

            if h.internal_value.lower() == key.lower():
                data = person['output'][key]

                # minus one because the Excel row number is included in the
                #  header cell list
                to_write[col - 1] = unique 
    
    # appending the updated list to the worksheet object, and then using
    #  a try-block to catch any writing errors
    ws.append(to_write)
    try:
        wb.save(output_path)
        log.emit("<br><br><b>Saved the results for " + person['name'] + \
                 " to Excel output.</b><br>")
        return True
    except PermissionError:
        log.emit("<br><br><b>An error was encountered when accessing the" \
                 " output Excel file. This likely means that it is open" \
                 " somewhere. Stoping scraping.</b>")
        return False


def main(input_path, output_name, output_format, log, table):
    """
    For some csv formatted correctly (ie has a header and is filled with
    researchers, their institutions, and their domains) this will get
    good links, then use analysis.py to gather the webtext and have gpt
    analyze it.

    At the end, a single researcher dict will have the following
      name: ...,
      institution: ...,
      any other elements included in the csv file: ...,
      links used: all the good links found by get_links(),
      output: the output created by analysis.analyze(), as a sub-dict.
        Note that the links used are also within this sub-dict.

    A loop will collect all of this information for each person; at the end
    of each loop, the Excel file will be updated for the new person. Updating
    continuously allows for data to be saved in case of any issues.
    
    input_path: the path to the input csv, gotten by interacting with the GUI
    output_name: the name/path of the output Excel that will be created. If
     there is already an Excel file with that path, it will be overwritten.
    log: a pyqtSignal(str) object that is used to update the GUI's log.
    table: a pyqtSignal(str) object that is used to update the GUI's table.
    """

    # gets a list of dicts built using the input path. If name/institution is
    #  not included in the input file, then scraping will be halted
    to_search = read_csv(input_path, table, log)
    if not to_search:
        return False

    # the below chunk updates the log with initial data: how many individuals,
    #  the output header to be used, the prompts to be used, and the sites
    #  identified by the user as good
    sing_or_plur = " individuals"
    if len(to_search) <= 1:
        sing_or_plur = " individual"
    log.emit("<h2>Starting scraping on " + str(len(to_search)) + \
             sing_or_plur + ".</h2><br><br><b>OUTPUT HEADER:</b><br>" + \
             ", ".join(output_format['headers']) + "<br>")
    for i in range(len(output_format['prompts'])):
        log.emit("<br><b>PROMPT " + str(i + 1) + "</b><br>")
        log.emit(output_format['prompts'][i] + "<br>")
    if len(output_format['sites']) > 0:
        log.emit("<br><b>ADDITIONAL SEARCH TERMS:</b><br>" + \
                 ", ".join(output_format['sites']) + "<br>")

    # attempts to build the Excel output file. build_good is a bool to show if
    #  the sheet was accessed okay: True if there where no issues and False if
    #  so. If False, then scraping will stop and the user will be notified
    build_good = build_output_file(output_name, output_format['headers'], log)
    if not build_good:
        return False

    # an instance of an OpenAI() object
    client = analysis.animate_client()
    log.emit("<br>OpenAI client successfully activated.<br>")

    # if the user wants emails to be found, then regex is used instead of
    #  openai. This bool ensures that happens
    get_email = False
    if "email" in output_format['headers'] \
       or "Email" in output_format['headers'] \
       or "emails" in output_format['headers'] \
       or "Emails" in output_format['headers']:
        log.emit("Regex will be used to scrape emails.<br>")
        get_email = True

    # this count is used to update the table in the GUI
    count = 0
    # the total time taken is displayed at the end
    total_time_start = time.time()
    for person in to_search:
        # time per person being scraped is displayed
        start = time.time()
    
        log.emit("<br><h3>Scraping " + person['name'] + ", " + \
                 person["institution"] + "</h3><br><br>")

        # the agent used for requests from the web
        agent = random.choice(user_agents)

        # good links found by get_links are added to each person dict
        #  google's programmable search is used to search google
        google_api_key = os.getenv("GOOGLE_CUSTOM_SEARCH_API_KEY")
        # this is the id of the programmable search engine in use
        cse_id = os.getenv("CSE_ID")
        person['links used'] = get_links(
            person, output_format['sites'], google_api_key, cse_id, log
        )

        # the log will update with any good links found, if any
        log.emit("<b>Sites found: ")
        if len(person['links used']) == 0:
            log.emit("N/A</b><br>")
        else:
            log.emit(str(len(person['links used'])) + "</b><br>")
            log.emit("<br>".join([f"<a href=\"{site}\">{site}</a>" \
                                  for site in person["links used"]]))

        # this gets all relevant found information for one person
        person['output'] = analysis.analyze(
            person, client, output_format['prompts'], output_format['headers'],
            get_email, log
        )

        # writes the found information to excel. the bool of build_good is
        #  exactly like that from build_output_file
        build_good = write_to_excel(output_name, person, log)
        if not build_good:
            return False

        # outputting final tidbits of information
        end = time.time()
        log.emit("<b>Time spent: " + str(round(end - start, 2)) + " s</b>")
        table.emit("completed:" + str(count))
        count += 1

    # outputting the total time taken, in minutes
    total_time_end = time.time()
    total_time = round((total_time_end - total_time_start) / 60, 2)
    log.emit("<br><br><b>Scraping complete. Total time spent: " + \
             str(total_time) + " minutes</b>")
